import functools

from django.contrib import messages
from django.core.cache import cache
from django.core.files.storage import FileSystemStorage
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.utils.datastructures import MultiValueDictKeyError
from openpyxl import load_workbook

from api.models import Device
from .admin import create_password
from .forms import TeacherLoginForm, PostgraduateLoginForm, GroupTeacherMemberForm, TeacherForm, PostgraduateForm
from .models import Teacher, Postgraduate, Group
from .verification import verify_teacher_by_password, verify_postgraduate_by_password

UPLOAD_XLSX_FILE = "import_data.xlsx"


def login_required(func):
    @functools.wraps(func)
    def wrapper(request, *args, **kw):
        if request.session.has_key('user') and request.session['user']:
            return func(request, *args, **kw)
        else:
            return redirect('login')

    return wrapper


def login(request):
    """教师、研究生登陆逻辑"""
    response = dict()
    # 教师登陆
    if request.path == reverse('teacher_login'):
        response['url'] = reverse('teacher_login')
        response['user_type'] = '教师'
        if request.method == 'POST':
            form = TeacherLoginForm(request.POST)
            if form.is_valid():
                teacher = verify_teacher_by_password(form.cleaned_data['phone'], form.cleaned_data['password'])
                if teacher:
                    request.session['type'] = 'teacher'
                    request.session['user'] = teacher.phone
                    return redirect('teacher_home')
                else:
                    form.add_error(None, '密码错误')
        else:
            form = TeacherLoginForm()
        response['form'] = form
    # 研究生登陆
    if request.path == reverse('postgraduate_login'):
        response['url'] = reverse('postgraduate_login')
        response['user_type'] = '研究生'
        if request.method == 'POST':
            form = PostgraduateLoginForm(request.POST)
            if form.is_valid():
                postgraduate = verify_postgraduate_by_password(form.cleaned_data["phone"],
                                                               form.cleaned_data['password'])
                if postgraduate:
                    request.session['type'] = 'postgraduate'
                    request.session['user'] = postgraduate.phone
                    return redirect('postgraduate_home')
                else:
                    form.add_error(None, '密码错误')
        else:
            form = PostgraduateLoginForm()
        response['form'] = form
    return render(request, 'account/login.html', response)


def logout(request):
    """教师、研究生登出逻辑"""
    try:
        del request.session['type']
        del request.session['user']
    except KeyError:
        pass
    return redirect('home')


def get_login_user(request):
    """返回已登陆的用户实例"""
    user_type = request.session.get('type')
    phone = request.session.get('user')
    if user_type == 'teacher':
        return Teacher.objects.get(phone=phone)
    if user_type == 'postgraduate':
        return Postgraduate.objects.get(phone=phone)


def home(request):
    """用户已登陆则跳转至对应主页"""
    login_user = get_login_user(request)
    if login_user is not None:
        if isinstance(login_user, Teacher):
            return redirect('teacher_home')
        if isinstance(login_user, Postgraduate):
            return redirect('postgraduate_home')
    return render(request, 'account/home.html')


@login_required
def teacher_home(request):
    response = dict()
    response['teacher'] = teacher = get_login_user(request)
    key = str(teacher.uuid) + '_invited'
    group_id = cache.get(key)
    if group_id and Group.objects.filter(id=group_id).exists():
        response['invited'] = Group.objects.get(id=group_id)
    return render(request, 'account/home_teacher.html', response)


@login_required
def members(request):
    response = dict()
    response['teacher'] = teacher = get_login_user(request)
    response['members'] = Teacher.objects.filter(group__leader=teacher).all()
    return render(request, 'account/members.html', response)


@login_required
def invite(request):
    response = dict()
    response['teacher'] = teacher = get_login_user(request)
    if request.method == 'POST':
        form = GroupTeacherMemberForm(data=request.POST, teacher=teacher)
        if form.is_valid():
            group = Group.objects.get(leader=teacher)
            for t in form.cleaned_data['teacher_member']:
                key = str(t.uuid) + '_invited'
                cache.set(key, group.id, None)
            messages.add_message(request, messages.INFO, '邀请发送成功')
            return redirect('teacher_home')
    else:
        form = GroupTeacherMemberForm(teacher=teacher)
    response['form'] = form
    return render(request, 'account/invite.html', response)


@login_required
def handle_invite(request, group_id):
    teacher = get_login_user(request)
    if request.method == 'POST':
        group = Group.objects.get(id=group_id)
        result = request.POST.get('result')
        if result == 'approve':
            teacher.group = group
            teacher.save()
            messages.add_message(request, messages.INFO, '已加入{}课题组'.format(group.name))
        elif result == 'reject':
            messages.add_message(request, messages.INFO, '已拒绝{}的邀请'.format(group.leader.name))
        key = str(teacher.uuid) + '_invited'
        cache.delete(key)
        return redirect('teacher_home')


@login_required
def remove(request):
    if request.method == 'GET':
        uuid = request.GET.get('uuid')
        if uuid:
            teacher = Teacher.objects.get(uuid=uuid)
            teacher.group = None
            teacher.save()
            messages.add_message(request, messages.INFO, '{} 已被移出组'.format(teacher.name))
    return redirect('g_members')


@login_required
def postgraduate_list(request):
    teacher = get_login_user(request)
    return render(request, 'account/postgraduate_list.html', {'teacher': teacher})


@login_required
def table_postgraduate_list(request):
    if request.method == 'GET':
        limit = int(request.GET.get('limit'))
        offset = int(request.GET.get('offset'))
        search = request.GET.get('search')
        sort_column = request.GET.get('sort')
        order = request.GET.get('order')
        teacher = get_login_user(request)
        if search:
            postgraduates = teacher.postgraduate_set.filter(Q(phone=search) | Q(name=search))  # 或查询需要试用django Q
        else:
            postgraduates = teacher.postgraduate_set.all()

        if sort_column:
            sort_column = sort_column.replace('postgraduate_', '')
            if sort_column in ['phone', 'name', 'teacher', 'school', 'classes']:
                if order == 'desc':
                    sort_column = '-{}'.format(sort_column)
                postgraduates = postgraduates.order_by(sort_column)

        response = {'total': postgraduates.count(), 'rows': []}
        for postgraduate in postgraduates:
            device = '未绑定'
            if Device.objects.filter(postgraduate=postgraduate).exists():
                if Device.objects.get(postgraduate=postgraduate).imei:
                    device = '已绑定'
            response['rows'].append({
                "postgraduate_phone": postgraduate.phone,
                "postgraduate_name": postgraduate.name,
                "postgraduate_teacher": postgraduate.teacher.name,
                "postgraduate_school": postgraduate.school,
                "postgraduate_classes": postgraduate.classes,
                "postgraduate_device": device
            })
        if not offset:
            offset = 0
        if not limit:
            limit = 20
        response['rows'] = response['rows'][offset:offset + limit]
        return JsonResponse(response)


@login_required
def add_postgraduate(request):
    response = dict()
    response['teacher'] = teacher = get_login_user(request)
    if request.method == 'POST':
        form = PostgraduateForm(request.POST)
        if form.is_valid():
            postgraduate = form.save(commit=False)
            postgraduate.teacher = teacher
            create_password(postgraduate)
            postgraduate.save()
            messages.add_message(request, messages.INFO, '添加成功')
            return redirect('postgraduate_list')
    else:
        form = PostgraduateForm()
    response['form'] = form
    return render(request, 'account/add_postgraduate.html', response)


@login_required
def import_postgraduate_list(request):
    teacher = get_login_user(request)
    response = {'teacher': teacher}
    try:
        if request.method == 'POST' and request.FILES['excel']:
            excel = request.FILES['excel']
            fs = FileSystemStorage()
            fs.delete(UPLOAD_XLSX_FILE)  # 删除暂存文件
            fs.save(UPLOAD_XLSX_FILE, excel)  # 暂存media文件夹中
            response['upload_file'] = True
            # 检查文件
            workbook = load_workbook(fs.path(UPLOAD_XLSX_FILE))
            sheet_names = workbook.get_sheet_names()
            worksheet = workbook.get_sheet_by_name(sheet_names[0])
            rows = worksheet.rows
            waring = []
            for row in rows:
                line = [col.value for col in row]
                if len(line) != 4:
                    response['upload_file'] = False
                    messages.add_message(request, messages.ERROR, '导入文件格式错误，请参照模板')
                    break
                if Postgraduate.objects.filter(phone=line[0]).exists():
                    waring.append(line[0])
                    continue
            if len(waring) > 0:
                messages.add_message(request, messages.WARNING, ','.join(waring) + '已在数据库中，将不会导入')

    except MultiValueDictKeyError:
        return redirect('import_postgraduate_list')
    if request.method == 'GET' and request.GET.get('confirm') == 'true':
        fs = FileSystemStorage()
        workbook = load_workbook(fs.path(UPLOAD_XLSX_FILE))
        sheet_names = workbook.get_sheet_names()
        worksheet = workbook.get_sheet_by_name(sheet_names[0])
        rows = worksheet.rows
        postgraduates = []
        for row in rows:
            line = [col.value for col in row]
            if line[0] == "手机号(必要)":
                continue  # 跳过标题
            if Postgraduate.objects.filter(phone=line[0]).exists():
                # 跳过已存在的数据
                continue
            postgraduate = Postgraduate(phone=line[0],
                                        password=str(line[0]),
                                        name=line[1],
                                        teacher=teacher,
                                        school=line[2],
                                        classes=line[3])
            create_password(postgraduate)
            postgraduates.append(postgraduate)
        Postgraduate.objects.bulk_create(postgraduates)
        fs.delete(UPLOAD_XLSX_FILE)  # 删除暂存文件
        return redirect('postgraduate_list')
    return render(request, 'account/import_postgraduate_list.html', response)


@login_required
def table_uploaded_postgraduate_list(request):
    if request.method == 'GET':
        limit = int(request.GET.get('limit'))
        offset = int(request.GET.get('offset'))
        fs = FileSystemStorage()
        workbook = load_workbook(fs.path(UPLOAD_XLSX_FILE))
        sheet_names = workbook.get_sheet_names()
        worksheet = workbook.get_sheet_by_name(sheet_names[0])
        rows = worksheet.rows
        response = dict()
        response['rows'] = []
        for row in rows:
            line = [col.value for col in row]
            if line[0] == "手机号(必要)":
                continue
            if Postgraduate.objects.filter(phone=line[0]).exists():
                continue
            response['rows'].append({
                "postgraduate_phone": line[0],
                "postgraduate_name": line[1],
                "postgraduate_school": line[2],
                "postgraduate_classes": line[3]
            })
        if not offset:
            offset = 0
        if not limit:
            limit = 20
        response['total'] = len(response['rows'])
        response['rows'] = response['rows'][offset:offset + limit]
        return JsonResponse(response)


@login_required
def import_teacher(request):
    response = dict()
    response['teacher'] = teacher = get_login_user(request)
    group = Group.objects.get(leader=teacher)
    try:
        if request.method == 'POST' and request.FILES['excel']:
            excel = request.FILES['excel']
            fs = FileSystemStorage()
            fs.delete(UPLOAD_XLSX_FILE)  # 删除暂存文件
            fs.save(UPLOAD_XLSX_FILE, excel)  # 暂存media文件夹中
            response['upload_file'] = True
            # 检查文件
            workbook = load_workbook(fs.path(UPLOAD_XLSX_FILE))
            sheet_names = workbook.get_sheet_names()
            worksheet = workbook.get_sheet_by_name(sheet_names[0])
            rows = worksheet.rows
            waring = []
            for row in rows:
                line = [col.value for col in row]
                if len(line) != 4:
                    response['upload_file'] = False
                    messages.add_message(request, messages.ERROR, '导入文件格式错误，请参照模板')
                    break
                if Teacher.objects.filter(phone=line[0]).exists():
                    waring.append(line[0])
                    continue
            if len(waring) > 0:
                messages.add_message(request, messages.WARNING, ', '.join(waring) + ' 已在数据库中，将不会导入')
    except MultiValueDictKeyError:
        return redirect('import_teacher')
    if request.method == 'GET' and request.GET.get('confirm') == 'true':
        fs = FileSystemStorage()
        workbook = load_workbook(fs.path(UPLOAD_XLSX_FILE))
        sheet_names = workbook.get_sheet_names()
        worksheet = workbook.get_sheet_by_name(sheet_names[0])
        rows = worksheet.rows
        teachers = []
        for row in rows:
            line = [col.value for col in row]
            if line[0] == '手机号(必要)':
                continue  # 跳过标题
            if Teacher.objects.filter(phone=line[0]).exists():
                continue
            t = Teacher(phone=line[0], password=str(line[0]), name=line[1], school=line[2], specialty=line[3])
            create_password(t)
            teachers.append(t)
        Teacher.objects.bulk_create(teachers)
        fs.delete(UPLOAD_XLSX_FILE)  # 删除暂存文件
        return HttpResponse('OK')
    return render(request, 'account/import_teacher_list.html', response)


@login_required
def table_uploaded_teacher_list(request):
    if request.method == 'GET':
        limit = int(request.GET.get('limit'))
        offset = int(request.GET.get('offset'))
        fs = FileSystemStorage()
        workbook = load_workbook(fs.path(UPLOAD_XLSX_FILE))
        sheet_names = workbook.get_sheet_names()
        worksheet = workbook.get_sheet_by_name(sheet_names[0])
        rows = worksheet.rows
        response = dict()
        response['rows'] = []
        for row in rows:
            line = [col.value for col in row]
            if line[0] == '手机号(必要)':
                continue
            if Teacher.objects.filter(phone=line[0]).exists():
                continue
            response['rows'].append({
                'teacher_phone': line[0],
                'teacher_name': line[1],
                'teacher_school': line[2],
                'teacher_specialty': line[3]

            })
        if not offset:
            offset = 0
        if not limit:
            limit = 20
        response['total'] = len(response['rows'])
        response['rows'] = response['rows'][offset:offset + limit]
        return JsonResponse(response)


@login_required
def postgraduate_home(request):
    postgraduate = get_login_user(request)
    return render(request, 'account/home_postgraduate.html', {'postgraduate': postgraduate})


@login_required
def setting_t(request):
    response = dict()
    response['teacher'] = teacher = get_login_user(request)
    if request.method == 'POST':
        form = TeacherForm(request.POST, instance=teacher)
        password = request.POST.get('password')
        password_new = request.POST.get('password1')
        if verify_teacher_by_password(teacher.phone, password):
            if form.is_valid():
                if form.has_changed():
                    form.save()
                    messages.add_message(request, messages.INFO, '个人资料修改成功')
            if password_new:
                teacher.password = password_new
                create_password(teacher)
                teacher.save()
                messages.add_message(request, messages.INFO, '密码修改成功')
        else:
            messages.add_message(request, messages.WARNING, '密码错误')
    else:
        form = TeacherForm(instance=teacher)
    response['form'] = form
    return render(request, 'account/setting.html', response)


@login_required
def setting_p(request):
    response = dict()
    response['postgraduate'] = postgraduate = get_login_user(request)
    if request.method == 'POST':
        form = PostgraduateForm(request.POST, instance=postgraduate)
        password = request.POST.get('password')
        password_new = request.POST.get('password1')
        if verify_teacher_by_password(postgraduate.phone, password):
            if form.is_valid():
                if form.has_changed():
                    form.save()
                    messages.add_message(request, messages.INFO, '个人资料修改成功')
            if password_new:
                postgraduate.password = password_new
                create_password(postgraduate)
                postgraduate.save()
                messages.add_message(request, messages.INFO, '密码修改成功')
        else:
            messages.add_message(request, messages.WARNING, '密码错误')
    else:
        form = PostgraduateForm(instance=postgraduate)
    response['form'] = form
    return render(request, 'account/setting.html', response)
