from datetime import datetime
from django.core.files.storage import FileSystemStorage
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from openpyxl import load_workbook

from management.forms import TeacherLoginForm, PostgraduateLoginForm
from management.models import Teacher, Postgraduate, CheckIn, Leave

UPLOAD_XLSX_FILE = "import_data.xlsx"


def home(request):
    """用户已登陆则跳转至对应主页"""
    login_user = __get_login_user(request)
    if login_user is not None:
        if isinstance(login_user, Teacher):
            return redirect('teacher_home')
        if isinstance(login_user, Postgraduate):
            return redirect('postgraduate_home')
    return render(request, 'home.html')


def choose_login_type(request):
    return render(request, 'choose_login_type.html')


def logout(request):
    try:
        del request.session['type']
        del request.session['user']
    except KeyError:
        pass
    return redirect('home')


def login(request):
    """教师、研究生登陆逻辑"""
    data = {}
    # 教师登陆
    if request.path == reverse('teacher_login'):
        data['url'] = reverse('teacher_login')
        data['user_type'] = '教师'
        if request.method == 'GET':
            form = TeacherLoginForm()
        else:
            form = TeacherLoginForm(request.POST)
            if form.is_valid():
                username = request.POST.get('username', '')
                password = request.POST.get('password', '')
                teacher = Teacher.objects.get(username=username)
                if teacher.password == password:
                    request.session['type'] = 'teacher'
                    request.session['user'] = teacher.username
                    return redirect('teacher_home')
        data['form'] = form
    # 研究生登陆
    if request.path == reverse('postgraduate_login'):
        data['url'] = reverse('postgraduate_login')
        data['user_type'] = '研究生'
        if request.method == 'GET':
            form = PostgraduateLoginForm()
        else:
            form = PostgraduateLoginForm(request.POST)
            if form.is_valid():
                _id = request.POST.get('id', '')
                password = request.POST.get('password', '')
                postgraduate = Postgraduate.objects.get(id=_id)
                if postgraduate.password == password:
                    request.session['type'] = 'postgraduate'
                    request.session['user'] = postgraduate.id
                    return redirect('postgraduate_home')
        data['form'] = form
    return render(request, 'login.html', data)


def teacher_home(request):
    teacher = __get_login_user(request)
    return render(request, 'home_teacher.html', {'teacher': teacher})


def postgraduate_list(request):
    teacher = __get_login_user(request)
    return render(request, 'postgraduate_list.html', {'teacher': teacher})


def table_postgraduate_list(request):
    if request.method == 'GET':
        limit = int(request.GET.get('limit'))
        offset = int(request.GET.get('offset'))
        search = request.GET.get('search')
        sort_column = request.GET.get('sort')
        order = request.GET.get('order')
        teacher = __get_login_user(request)
        if search:
            postgraduates = teacher.postgraduate_set.filter(Q(id=search) | Q(name=search))  # 或查询需要试用django Q
        else:
            postgraduates = teacher.postgraduate_set.all()

        if sort_column:
            sort_column = sort_column.replace('postgraduate_', '')
            if sort_column in ['id', 'name', 'teacher', 'group']:
                if order == 'desc':
                    sort_column = '-{}'.format(sort_column)
                postgraduates = postgraduates.order_by(sort_column)

        response_data = {'total': postgraduates.count(), 'rows': []}
        for postgraduate in postgraduates:
            response_data['rows'].append({
                "postgraduate_id": postgraduate.id,
                "postgraduate_name": postgraduate.name,
                "postgraduate_teacher": postgraduate.teacher.username,
                "postgraduate_group": postgraduate.group if postgraduate.group else "",
            })

        if not offset:
            offset = 0
        if not limit:
            limit = 20
        response_data['rows'] = response_data['rows'][offset:offset + limit]
        return JsonResponse(response_data)


def import_postgraduate_list(request):
    teacher = __get_login_user(request)
    response_data = {'teacher': teacher}
    if request.method == 'POST' and request.FILES['excel']:
        excel = request.FILES['excel']
        fs = FileSystemStorage()
        fs.save(UPLOAD_XLSX_FILE, excel)  # 暂存media文件夹中
        response_data['upload_file'] = True
    elif request.method == 'GET' and request.GET.get('confirm') == 'true':
        fs = FileSystemStorage()
        workbook = load_workbook(fs.path(UPLOAD_XLSX_FILE))
        sheet_names = workbook.get_sheet_names()
        worksheet = workbook.get_sheet_by_name(sheet_names[0])
        rows = worksheet.rows
        postgraduates = []
        for row in rows:
            line = [col.value for col in row]
            if line[0] == "学号":
                continue  # 跳过标题（TODO：以是否为数字作为判断）
            postgraduates.append(Postgraduate(id=line[0],
                                              password='123',
                                              name=line[1],
                                              teacher=teacher))
        Postgraduate.objects.bulk_create(postgraduates)
        fs.delete(UPLOAD_XLSX_FILE)  # 删除暂存文件
        return HttpResponse('OK')
    return render(request, 'import_postgraduate_list.html', response_data)


def table_uploaded_postgraduate_list(request):
    if request.method == 'GET':
        limit = int(request.GET.get('limit'))
        offset = int(request.GET.get('offset'))

        fs = FileSystemStorage()
        workbook = load_workbook(fs.path(UPLOAD_XLSX_FILE))
        sheet_names = workbook.get_sheet_names()
        worksheet = workbook.get_sheet_by_name(sheet_names[0])
        rows = worksheet.rows
        response_data = {'total': 0, 'rows': []}
        for row in rows:
            line = [col.value for col in row]
            if line[0] == "学号":
                continue
            # noinspection PyTypeChecker
            response_data['rows'].append({
                "postgraduate_id": line[0],
                "postgraduate_name": line[1],
            })

        if not offset:
            offset = 0
        if not limit:
            limit = 20
        response_data['total'] = len(response_data['rows'])
        response_data['rows'] = response_data['rows'][offset:offset + limit]
        return JsonResponse(response_data)


def postgraduate_home(request):
    postgraduate = __get_login_user(request)
    return render(request, 'home_postgraduate.html', {'postgraduate': postgraduate})


def check_in(request):
    if request.method == 'GET':
        return render(request, 'check_in.html')
    else:
        postgraduate = Postgraduate.objects.get(id=request.POST.get('postgraduate'))
        records = CheckIn.objects.filter(date=datetime.now().date(), postgraduate=postgraduate)
        if records:
            record = records[0]
            changed = False
            current_time = datetime.now().time()
            if record.forenoon_out is None:
                record.forenoon_out = current_time
                changed = True
            elif record.afternoon_in is None:
                record.afternoon_in = current_time
                changed = True
            elif record.afternoon_out is None:
                record.afternoon_out = current_time
                changed = True
            if changed:
                record.save()
                return redirect('check_in')
            else:
                return HttpResponse('今日签到已完成！')
        else:
            record = CheckIn.objects.create(postgraduate=postgraduate, date=datetime.now().date())
        record.forenoon_in = datetime.now().time()
        record.save()
        return redirect('check_in')


def show_check_in(request):
    teacher = __get_login_user(request)
    response_data = {'teacher': teacher}
    response_data['startDate'] = datetime.now().date().strftime("%Y-%m-%d")
    return render(request, 'show_check_in.html', response_data)


def ask_for_leave(request):
    postgraduate = __get_login_user(request)
    response_data = {'postgraduate': postgraduate}
    if request.method == 'POST':
        date = request.POST.get('date')
        excuse = request.POST.get('excuse')
        date = datetime.strptime(date, '%Y-%m-%d').date()
        # TODO:是否重复
        Leave.objects.create(
            postgraduate=postgraduate,
            date=date,
            excuse=excuse,
            time_of_submission=datetime.now()
        )
        return HttpResponse('提交成功！')
    return render(request, 'ask_for_leave.html', response_data)


def leave_list(request):
    teacher = __get_login_user(request)
    response_data = {'teacher': teacher}
    postgraduates = teacher.postgraduate_set.all()
    leaves = Leave.objects.filter(state=None).all()
    leave_set = set()
    for leave in leaves:
        if leave.postgraduate in postgraduates:
            leave_set.add(leave)
    response_data['leave_set'] = leave_set
    return render(request, 'leave_list.html', response_data)


def __get_login_user(request):
    """返回已登陆的用户实例"""
    user_type = request.session.get('type')
    if user_type == 'teacher':
        username = request.session.get('user')
        return Teacher.objects.get(username=username)
    if user_type == 'postgraduate':
        _id = request.session.get('user')
        return Postgraduate.objects.get(id=_id)
